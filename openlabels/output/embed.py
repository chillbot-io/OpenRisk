"""
OpenLabels Embedded Label Writer.

Writes labels directly into file native metadata:
- PDF: XMP metadata
- DOCX/XLSX/PPTX: Custom Document Properties
- Images (JPEG/PNG/TIFF): XMP or EXIF

Per the spec, embedded labels contain the full LabelSet JSON and are
the source of truth for files that support native metadata.
"""

import logging
from pathlib import Path
from typing import Optional, Union
from abc import ABC, abstractmethod

from ..core.labels import LabelSet

logger = logging.getLogger(__name__)

# XMP namespace for OpenLabels
OPENLABELS_XMP_NS = "http://openlabels.dev/ns/1.0/"
OPENLABELS_XMP_PREFIX = "openlabels"


class EmbeddedLabelWriter(ABC):
    """Abstract base class for embedded label writers."""

    @abstractmethod
    def write(self, path: Path, label_set: LabelSet) -> bool:
        """
        Write a LabelSet to the file's native metadata.

        Args:
            path: Path to the file
            label_set: The LabelSet to embed

        Returns:
            True if successful, False otherwise
        """
        pass

    @abstractmethod
    def read(self, path: Path) -> Optional[LabelSet]:
        """
        Read a LabelSet from the file's native metadata.

        Args:
            path: Path to the file

        Returns:
            LabelSet if found, None otherwise
        """
        pass

    @abstractmethod
    def supports(self, path: Path) -> bool:
        """Check if this writer supports the given file type."""
        pass


class PDFLabelWriter(EmbeddedLabelWriter):
    """
    Write/read labels to PDF XMP metadata.

    Uses pikepdf for PDF manipulation.
    Stores in XMP namespace: http://openlabels.dev/ns/1.0/
    Property name: openlabels
    """

    SUPPORTED_EXTENSIONS = {'.pdf'}

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def write(self, path: Path, label_set: LabelSet) -> bool:
        """Write LabelSet to PDF XMP metadata."""
        try:
            import pikepdf
        except ImportError:
            logger.warning("pikepdf not installed, cannot write PDF labels")
            return False

        try:
            with pikepdf.open(path, allow_overwriting_input=True) as pdf:
                with pdf.open_metadata() as meta:
                    meta[f'{{{OPENLABELS_XMP_NS}}}openlabels'] = label_set.to_json(compact=True)
                pdf.save(path)
            return True

        except (OSError, ValueError) as e:
            logger.error(f"Failed to write PDF labels: {e}")
            return False

    def read(self, path: Path) -> Optional[LabelSet]:
        """Read LabelSet from PDF XMP metadata."""
        try:
            import pikepdf
        except ImportError:
            logger.warning("pikepdf not installed, cannot read PDF labels")
            return None

        try:
            with pikepdf.open(path) as pdf:
                with pdf.open_metadata() as meta:
                    # Try to read our namespace
                    key = f'{{{OPENLABELS_XMP_NS}}}openlabels'
                    if key in meta:
                        json_str = str(meta[key])
                        return LabelSet.from_json(json_str)
            return None

        except (OSError, ValueError) as e:
            logger.debug(f"No labels found in PDF: {e}")
            return None


class OfficeLabelWriter(EmbeddedLabelWriter):
    """
    Write/read labels to Office document custom properties.

    Supports DOCX, XLSX, PPTX (all are ZIP-based OOXML formats).
    Stores in Custom Document Properties with name "openlabels".
    """

    SUPPORTED_EXTENSIONS = {'.docx', '.xlsx', '.pptx'}

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def write(self, path: Path, label_set: LabelSet) -> bool:
        """Write LabelSet to Office custom properties."""
        suffix = path.suffix.lower()

        try:
            if suffix == '.docx':
                return self._write_docx(path, label_set)
            elif suffix == '.xlsx':
                return self._write_xlsx(path, label_set)
            elif suffix == '.pptx':
                return self._write_pptx(path, label_set)
            return False
        except (OSError, ValueError, KeyError) as e:
            logger.error(f"Failed to write Office labels: {e}")
            return False

    def read(self, path: Path) -> Optional[LabelSet]:
        """Read LabelSet from Office custom properties."""
        suffix = path.suffix.lower()

        try:
            if suffix == '.docx':
                return self._read_docx(path)
            elif suffix == '.xlsx':
                return self._read_xlsx(path)
            elif suffix == '.pptx':
                return self._read_pptx(path)
            return None
        except (OSError, ValueError, KeyError) as e:
            logger.debug(f"No labels found in Office doc: {e}")
            return None

    def _write_docx(self, path: Path, label_set: LabelSet) -> bool:
        """Write to DOCX custom properties."""
        try:
            from docx import Document
        except ImportError:
            logger.warning("python-docx not installed, cannot write DOCX labels")
            return False

        doc = Document(path)
        core_props = doc.core_properties
        # Use custom_properties if available, otherwise fall back to comments
        # python-docx doesn't have great custom property support,
        # so we'll use the comments field as a workaround
        core_props.comments = label_set.to_json(compact=True)
        doc.save(path)
        return True

    def _read_docx(self, path: Path) -> Optional[LabelSet]:
        """Read from DOCX custom properties."""
        try:
            from docx import Document
        except ImportError:
            return None

        doc = Document(path)
        comments = doc.core_properties.comments
        if comments and comments.startswith('{"v":'):
            return LabelSet.from_json(comments)
        return None

    def _write_xlsx(self, path: Path, label_set: LabelSet) -> bool:
        """Write to XLSX custom properties."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl not installed, cannot write XLSX labels")
            return False

        wb = load_workbook(path)
        # Use custom doc properties
        if wb.properties is None:
            from openpyxl.packaging.core import DocumentProperties
            wb.properties = DocumentProperties()
        wb.properties.description = label_set.to_json(compact=True)
        wb.save(path)
        return True

    def _read_xlsx(self, path: Path) -> Optional[LabelSet]:
        """Read from XLSX custom properties."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            return None

        wb = load_workbook(path, read_only=True)
        if wb.properties and wb.properties.description:
            desc = wb.properties.description
            if desc.startswith('{"v":'):
                return LabelSet.from_json(desc)
        return None

    def _write_pptx(self, path: Path, label_set: LabelSet) -> bool:
        """Write to PPTX custom properties."""
        try:
            from pptx import Presentation
        except ImportError:
            logger.warning("python-pptx not installed, cannot write PPTX labels")
            return False

        prs = Presentation(path)
        prs.core_properties.comments = label_set.to_json(compact=True)
        prs.save(path)
        return True

    def _read_pptx(self, path: Path) -> Optional[LabelSet]:
        """Read from PPTX custom properties."""
        try:
            from pptx import Presentation
        except ImportError:
            return None

        prs = Presentation(path)
        comments = prs.core_properties.comments
        if comments and comments.startswith('{"v":'):
            return LabelSet.from_json(comments)
        return None


class ImageLabelWriter(EmbeddedLabelWriter):
    """
    Write/read labels to image XMP/EXIF metadata.

    Supports JPEG, PNG, TIFF, WebP.
    Prefers XMP metadata, falls back to EXIF UserComment for JPEG.
    """

    SUPPORTED_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.tiff', '.tif', '.webp'}

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in self.SUPPORTED_EXTENSIONS

    def write(self, path: Path, label_set: LabelSet) -> bool:
        """Write LabelSet to image metadata."""
        try:
            # Try using piexif for JPEG EXIF
            if path.suffix.lower() in {'.jpg', '.jpeg'}:
                return self._write_jpeg_exif(path, label_set)

            # For other formats, try PIL with custom metadata
            return self._write_pil_metadata(path, label_set)

        except (OSError, ValueError) as e:
            logger.error(f"Failed to write image labels: {e}")
            return False

    def read(self, path: Path) -> Optional[LabelSet]:
        """Read LabelSet from image metadata."""
        try:
            if path.suffix.lower() in {'.jpg', '.jpeg'}:
                result = self._read_jpeg_exif(path)
                if result:
                    return result

            return self._read_pil_metadata(path)

        except (OSError, ValueError) as e:
            logger.debug(f"No labels found in image: {e}")
            return None

    def _write_jpeg_exif(self, path: Path, label_set: LabelSet) -> bool:
        """Write to JPEG EXIF UserComment."""
        try:
            import piexif
        except ImportError:
            logger.warning("piexif not installed, trying PIL fallback")
            return self._write_pil_metadata(path, label_set)

        try:
            exif_dict = piexif.load(str(path))
        except (OSError, ValueError) as e:
            logger.debug(f"Could not load existing EXIF from {path}, creating new: {e}")
            exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "1st": {}}

        # Store in UserComment (tag 37510)
        json_data = label_set.to_json(compact=True)
        # UserComment requires specific encoding
        user_comment = b'ASCII\x00\x00\x00' + json_data.encode('utf-8')
        exif_dict['Exif'][piexif.ExifIFD.UserComment] = user_comment

        exif_bytes = piexif.dump(exif_dict)
        piexif.insert(exif_bytes, str(path))
        return True

    def _read_jpeg_exif(self, path: Path) -> Optional[LabelSet]:
        """Read from JPEG EXIF UserComment."""
        try:
            import piexif
        except ImportError:
            return None

        try:
            exif_dict = piexif.load(str(path))
            user_comment = exif_dict.get('Exif', {}).get(piexif.ExifIFD.UserComment)
            if user_comment:
                # Strip encoding prefix (first 8 bytes)
                if user_comment.startswith(b'ASCII\x00\x00\x00'):
                    json_str = user_comment[8:].decode('utf-8')
                    if json_str.startswith('{"v":'):
                        return LabelSet.from_json(json_str)
        except (OSError, ValueError, KeyError) as e:
            logger.debug(f"Could not read EXIF label from {path}: {e}")
        return None

    def _write_pil_metadata(self, path: Path, label_set: LabelSet) -> bool:
        """Write using PIL with PNG text chunks."""
        try:
            from PIL import Image
            from PIL.PngImagePlugin import PngInfo
        except ImportError:
            logger.warning("PIL not installed, cannot write image labels")
            return False

        # SECURITY FIX (MED-006): Use try/finally to ensure PIL image is closed
        img = Image.open(path)
        try:
            suffix = path.suffix.lower()

            if suffix == '.png':
                # PNG supports text metadata natively
                metadata = PngInfo()
                metadata.add_text("openlabels", label_set.to_json(compact=True))

                # Preserve existing metadata
                if hasattr(img, 'info'):
                    for key, value in img.info.items():
                        if key != 'openlabels' and isinstance(value, str):
                            metadata.add_text(key, value)

                img.save(path, pnginfo=metadata)
                return True

            # For other formats, PIL doesn't have great metadata support
            logger.warning(f"Limited metadata support for {suffix}, label may not persist")
            return False
        finally:
            img.close()

    def _read_pil_metadata(self, path: Path) -> Optional[LabelSet]:
        """Read using PIL text chunks."""
        try:
            from PIL import Image
        except ImportError:
            return None

        try:
            # SECURITY FIX (MED-006): Use try/finally to ensure PIL image is closed
            img = Image.open(path)
            try:
                if hasattr(img, 'info') and 'openlabels' in img.info:
                    json_str = img.info['openlabels']
                    if json_str.startswith('{"v":'):
                        return LabelSet.from_json(json_str)
            finally:
                img.close()
        except (OSError, ValueError, KeyError) as e:
            logger.debug(f"Could not read PIL label from {path}: {e}")
        return None


# =============================================================================
# UNIFIED INTERFACE
# =============================================================================

# Registry of writers by file type
_WRITERS = [
    PDFLabelWriter(),
    OfficeLabelWriter(),
    ImageLabelWriter(),
]


def get_writer(path: Union[str, Path]) -> Optional[EmbeddedLabelWriter]:
    """Get the appropriate writer for a file type."""
    path = Path(path)
    for writer in _WRITERS:
        if writer.supports(path):
            return writer
    return None


def supports_embedded_labels(path: Union[str, Path]) -> bool:
    """Check if a file type supports embedded labels."""
    return get_writer(path) is not None


def write_embedded_label(
    path: Union[str, Path],
    label_set: LabelSet,
) -> bool:
    """
    Write a LabelSet to a file's native metadata.

    Args:
        path: Path to the file
        label_set: The LabelSet to embed

    Returns:
        True if successful, False otherwise

    Example:
        >>> from openlabels.core.labels import LabelSet, Label
        >>> labels = [Label(type="SSN", confidence=0.99, detector="checksum", value_hash="15e2b0")]
        >>> label_set = LabelSet.create(labels, content, source="openlabels:1.0.0")
        >>> write_embedded_label("document.pdf", label_set)
        True
    """
    path = Path(path)
    writer = get_writer(path)
    if writer is None:
        logger.warning(f"No embedded label writer for {path.suffix}")
        return False
    return writer.write(path, label_set)


def read_embedded_label(path: Union[str, Path]) -> Optional[LabelSet]:
    """
    Read a LabelSet from a file's native metadata.

    Args:
        path: Path to the file

    Returns:
        LabelSet if found, None otherwise

    Example:
        >>> label_set = read_embedded_label("document.pdf")
        >>> if label_set:
        ...     print(f"Found {len(label_set.labels)} labels")
    """
    path = Path(path)
    writer = get_writer(path)
    if writer is None:
        return None
    return writer.read(path)
